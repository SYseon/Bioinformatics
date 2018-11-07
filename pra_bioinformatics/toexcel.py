# -*- coding: utf-8 -*-
"""
Created on Tue Oct 30 18:14:03 2018

@author: leedongjae
"""

# import modules
import pandas as pd
import numpy as np
from openpyxl import load_workbook 
 
idx = 1
def excel_init(file_name):    
    
    excel_filename = file_name.split('/')[-1]
    # index_format(index) & columns_format(columns)정의
    columns_format = ['전처리기', '특징 선택', 'PCA 수', '분류기', 'KNN 개수','Accuracy', 'Precision 1', 'Precision 2', 'Recall 1', 'Recall 2', 'fbeta score 1', 'fbeta score 2', 'Support 1', 'Support 2', 'Confusion Mat[0,0]', 'Confusion Mat[0,1]','Confusion Mat[1,0]','Confusion Mat[1,1]']
    # DataFrame 초기화
    values = pd.DataFrame(columns=columns_format)
    # saves DataFrame(values) into an Excel file
    values.to_excel('./test.xlsx',
        sheet_name=excel_filename,
        columns = columns_format,
        header=True,
        startrow=0,
        startcol=0,
        engine=None,
        merge_cells=True,
        encoding=None,
        inf_rep='inf',
        verbose=True,
        freeze_panes=None
    )

def set_value(preprocessor_name, fselector_name, pca_num, classifier_name, knn_k, result_value, precision, recall, fbeta_score, support, conf_mat):
    global idx
    excel_filename = './test.xlsx'
    wb = load_workbook(filename = excel_filename)
    
    
    # get the worksheet
  
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    ws.cell(row=idx+1, column=1).value = idx
    ws.cell(row=idx+1, column=2).value = preprocessor_name
    ws.cell(row=idx+1, column=3).value = fselector_name
    ws.cell(row=idx+1, column=4).value = pca_num
    ws.cell(row=idx+1, column=5).value = classifier_name
    
    if knn_k is not None:
        ws.cell(row=idx+1, column=6).value = knn_k
        
    ws.cell(row=idx+1, column=7).value = result_value
    ws.cell(row=idx+1, column=8).value = precision[0]
    ws.cell(row=idx+1, column=9).value = precision[1]
    ws.cell(row=idx+1, column=10).value = recall[0]
    ws.cell(row=idx+1, column=11).value = recall[1]
    ws.cell(row=idx+1, column=12).value = fbeta_score[0]
    ws.cell(row=idx+1, column=13).value = fbeta_score[1]
    ws.cell(row=idx+1, column=14).value = support[0]
    ws.cell(row=idx+1, column=15).value = support[1]
    ws.cell(row=idx+1, column=16).value = conf_mat[0][0]
    ws.cell(row=idx+1, column=17).value = conf_mat[0][1]
    ws.cell(row=idx+1, column=18).value = conf_mat[1][0]
    ws.cell(row=idx+1, column=19).value = conf_mat[1][1]
    
    idx+=1
    
    wb.save(filename=excel_filename)
